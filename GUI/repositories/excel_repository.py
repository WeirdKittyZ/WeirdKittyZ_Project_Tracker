from pathlib import Path
from typing import Dict, List, Optional, Any
from openpyxl import Workbook, load_workbook

from utils.constants import (
    SHEETS,
    PROJECT_COLUMNS,
    TASK_COLUMNS,
    STATUS_HISTORY_COLUMNS,
    COUNTER_COLUMNS,
)


class ExcelRepository:
    def __init__(self, workbook_path: Path):
        self.workbook_path = Path(workbook_path)
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        if not self.workbook_path.exists():
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            sheet_definitions = {
                SHEETS["projects"]: PROJECT_COLUMNS,
                SHEETS["tasks"]: TASK_COLUMNS,
                SHEETS["status_history"]: STATUS_HISTORY_COLUMNS,
                SHEETS["counters"]: COUNTER_COLUMNS,
            }
            for sheet_name, columns in sheet_definitions.items():
                ws = wb.create_sheet(sheet_name)
                ws.append(columns)
            counters = wb[SHEETS["counters"]]
            for name in ["project", "task", "history"]:
                counters.append([name, 0])
            wb.save(self.workbook_path)
        else:
            self._validate_or_repair_workbook()

    def _validate_or_repair_workbook(self) -> None:
        wb = load_workbook(self.workbook_path)
        changed = False
        required = {
            SHEETS["projects"]: PROJECT_COLUMNS,
            SHEETS["tasks"]: TASK_COLUMNS,
            SHEETS["status_history"]: STATUS_HISTORY_COLUMNS,
            SHEETS["counters"]: COUNTER_COLUMNS,
        }
        for sheet_name, columns in required.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(columns)
                changed = True
            else:
                ws = wb[sheet_name]
                header = [cell.value for cell in ws[1]]
                if header != columns:
                    raise ValueError(
                        f"Invalid schema in sheet {sheet_name}. Expected {columns}, found {header}"
                    )

        # V4 no longer uses Activity_Log. If an older workbook is copied in,
        # remove that worksheet so the exported/opened Excel file stays clean.
        if "Activity_Log" in wb.sheetnames:
            del wb["Activity_Log"]
            changed = True
        counters = wb[SHEETS["counters"]]
        existing = {row[0] for row in counters.iter_rows(min_row=2, values_only=True) if row[0]}
        for name in ["project", "task", "history"]:
            if name not in existing:
                counters.append([name, 0])
                changed = True
        if changed:
            wb.save(self.workbook_path)

    def _load(self):
        return load_workbook(self.workbook_path)

    def read_table(self, sheet_name: str) -> List[Dict[str, Any]]:
        wb = self._load()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = list(rows[0])
        result = []
        for row in rows[1:]:
            if all(value is None for value in row):
                continue
            result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        return result

    def append_row(self, sheet_name: str, record: Dict[str, Any]) -> None:
        wb = self._load()
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        ws.append([record.get(header) for header in headers])
        wb.save(self.workbook_path)

    def update_row(self, sheet_name: str, id_column: str, id_value: str, updates: Dict[str, Any]) -> bool:
        wb = self._load()
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        id_idx = headers.index(id_column) + 1
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=id_idx).value == id_value:
                for key, value in updates.items():
                    if key in headers:
                        ws.cell(row=row_idx, column=headers.index(key) + 1).value = value
                wb.save(self.workbook_path)
                return True
        return False

    def get_by_id(self, sheet_name: str, id_column: str, id_value: str) -> Optional[Dict[str, Any]]:
        for record in self.read_table(sheet_name):
            if record.get(id_column) == id_value:
                return record
        return None

    def next_id(self, counter_name: str, prefix: str) -> str:
        wb = self._load()
        ws = wb[SHEETS["counters"]]
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == counter_name:
                current = int(ws.cell(row=row_idx, column=2).value or 0)
                new_value = current + 1
                ws.cell(row=row_idx, column=2).value = new_value
                wb.save(self.workbook_path)
                return f"{prefix}{new_value:05d}"
        ws.append([counter_name, 1])
        wb.save(self.workbook_path)
        return f"{prefix}00001"
